"""
excel_builder.py — Ultra Prompt v6.1
Construtor profissional de planilhas Excel (.xlsx)

Exports:
    - build_professional_xlsx(data, config) -> str

Dependencias: openpyxl
"""

__all__ = ["build_professional_xlsx"]

import os
import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Paletas de cores por tema
# ---------------------------------------------------------------------------
_THEMES = {
    "corporate": {
        "header_bg": "1F4E79",
        "header_font": "FFFFFF",
        "alt_row": "D6E4F0",
        "border": "8DB4E2",
    },
    "minimal": {
        "header_bg": "4A4A4A",
        "header_font": "FFFFFF",
        "alt_row": "F2F2F2",
        "border": "CCCCCC",
    },
    "vivid": {
        "header_bg": "A23B72",
        "header_font": "FFFFFF",
        "alt_row": "FDE8F0",
        "border": "E888B0",
    },
}


def _safe_output_path(path: str) -> str:
    """Gera um caminho seguro, sem sobrescrever arquivo existente."""
    if not os.path.exists(path):
        return path
    base, ext = os.path.splitext(path)
    counter = 1
    while os.path.exists(f"{base}_{counter}{ext}"):
        counter += 1
    return f"{base}_{counter}{ext}"


def build_professional_xlsx(
    data: List[Dict[str, Any]],
    config: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Gera uma planilha Excel profissional a partir de uma lista de dicionarios.

    Parametros
    ----------
    data : list[dict]
        Lista de dicionarios onde cada item e uma linha. As chaves do
        primeiro dicionario definem as colunas.
    config : dict, opcional
        Opcoes de configuracao:
            - headers        : list[str]  — nomes personalizados para colunas
            - color_theme    : str        — 'corporate' | 'minimal' | 'vivid' (padrao: 'corporate')
            - include_chart  : bool       — adiciona grafico de barras (padrao: False)
            - include_filters: bool       — ativa autofiltro (padrao: True)
            - include_totals : bool       — adiciona linha de totais SUM (padrao: False)
            - sheet_name     : str        — nome da aba (padrao: 'Dados')
            - output_path    : str        — caminho do arquivo de saida

    Retorno
    -------
    str
        Caminho absoluto do arquivo .xlsx gerado.
    """
    if not data:
        raise ValueError("A lista de dados esta vazia. Nenhuma planilha gerada.")

    config = config or {}
    theme_name = config.get("color_theme", "corporate")
    theme = _THEMES.get(theme_name, _THEMES["corporate"])
    sheet_name = config.get("sheet_name", "Dados")
    include_chart = config.get("include_chart", False)
    include_filters = config.get("include_filters", True)
    include_totals = config.get("include_totals", False)

    # Colunas e headers
    columns = list(data[0].keys())
    headers = config.get("headers", columns)

    # Caminho de saida
    default_name = f"relatorio_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = config.get("output_path", os.path.join("/home/vercel-sandbox", default_name))
    output_path = _safe_output_path(output_path)

    # Estilos
    header_fill = PatternFill(start_color=theme["header_bg"], end_color=theme["header_bg"], fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color=theme["header_font"], size=11)
    alt_fill = PatternFill(start_color=theme["alt_row"], end_color=theme["alt_row"], fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color=theme["border"]),
        right=Side(style="thin", color=theme["border"]),
        top=Side(style="thin", color=theme["border"]),
        bottom=Side(style="thin", color=theme["border"]),
    )
    data_font = Font(name="Calibri", size=11)
    num_format = "#,##0.00"

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # --- Header ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # --- Dados ---
    for row_idx, record in enumerate(data, start=2):
        for col_idx, key in enumerate(columns, start=1):
            value = record.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            # Formatacao numerica
            if isinstance(value, (int, float)):
                cell.number_format = num_format
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Cores alternadas
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    total_data_rows = len(data)
    total_cols = len(columns)

    # --- Linha de totais ---
    if include_totals:
        totals_row = total_data_rows + 2
        for col_idx, key in enumerate(columns, start=1):
            cell = ws.cell(row=totals_row, column=col_idx)
            # Verifica se a coluna tem dados numericos
            sample = data[0].get(key, "")
            if isinstance(sample, (int, float)):
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{total_data_rows + 1})"
                cell.number_format = num_format
            elif col_idx == 1:
                cell.value = "TOTAL"
            cell.font = Font(name="Calibri", bold=True, size=11, color=theme["header_font"])
            cell.border = thin_border
            cell.fill = header_fill

    # --- Freeze e autofiltro ---
    ws.freeze_panes = "A2"
    if include_filters:
        ws.auto_filter.ref = f"A1:{get_column_letter(total_cols)}{total_data_rows + 1}"

    # --- Largura automatica ---
    for col_idx in range(1, total_cols + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row_idx in range(2, total_data_rows + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # --- Grafico de barras ---
    if include_chart and total_data_rows > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = config.get("chart_title", sheet_name)
        chart.y_axis.title = "Valores"
        chart.x_axis.title = headers[0] if headers else "Categoria"

        # Usa a primeira coluna como categoria e colunas numericas como series
        cats = Reference(ws, min_col=1, min_row=2, max_row=total_data_rows + 1)
        for col_idx, key in enumerate(columns, start=1):
            sample = data[0].get(key, "")
            if isinstance(sample, (int, float)):
                values = Reference(ws, min_col=col_idx, min_row=1, max_row=total_data_rows + 1)
                chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, f"{get_column_letter(total_cols + 2)}2")

    # Salvar
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    return os.path.abspath(output_path)
